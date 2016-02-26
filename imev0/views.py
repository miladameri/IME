
# -*- coding: utf-8 -*-
from django.views.generic import TemplateView, View
from django.http import HttpResponse
import json
from datetime import timedelta
import datetime
import jdatetime as jdate
from openpyxl import load_workbook
from imev0.models import *
from django.db.models import Sum

ONE_WEEK = 1
THREE_WEEK = 3
THREE_MONTHS = 12
ONE_YEAR = 52


class Test(TemplateView):
    template_name = 'test.html'
    #template_name = 'test.html'

    def get_context_data(self, **kwargs):
        context = super(Test, self).get_context_data(**kwargs)
        return context


class Index(TemplateView):
    template_name = 'index.html'
    #template_name = 'test.html'

    def get_context_data(self, **kwargs):
        mains = []
        for x in MainGroup.objects.all():
            mains.append(str(x.name))
        context = {'datas': [1,2]}
        print(context)
        return context




class Datas(View):
    # def __init__(self):
    #     self.datas = []
    #     # self.product_producer = ['NCI-CCAA-00', 'NCI-CR08AB-00', 'NCI-SLG-00', 'NCI-SLR-00', 'CWD-CR08AB-00']
    #     self.product_producer = ['NCI-CR08AB-00']
    #     self.initialize()
    #
    # def get(self, request, *args, **kwargs):
    #     per = request.GET['select_and_time[date]']
    #     time_slot = 12
    #     code = {u"۰":"0",u"۱":"1",u"۲":"2",u"۳":"3",u"۴":"4",u"۵":"5",u"۶":"6",u"۷":"7",u"۸":"8",u"۹":"9", u"/":"/"}
    #     new_per = ''.join(code.get(ch, ch) for ch in per)
    #     dates = new_per.split('/')
    #
    #     if (dates == ['']):
    #         persian_date = jdate.date(1394, 8, 1)
    #     else:
    #         persian_date = jdate.date(int(dates[0]), int(dates[1]), int(dates[2]))
    #     print("dates:" , persian_date)
    #     result = self.get_product_producer('copper',persian_date, time_slot )
    #     datas = result
    #     output = {'datas': datas}
    #     print("datas",  datas)
    #     return HttpResponse(json.dumps(output))
    #
    #
    # def initialize(self):
    #     wb = load_workbook(filename='imev0/test2.xlsx', read_only=True)
    #     ws = wb.active  # ws is now an IterableWorksheet
    #     for row in ws:
    #         d = {}
    #
    #         if row[0].value == 'تاریخ':
    #             continue
    #         dates = row[0].value.split('/')
    #         persian_date = jdate.date(int(dates[0]), int(dates[1]), int(dates[2]))
    #
    #         d['date'] = persian_date
    #         d['name'] = row[1].value
    #         d['producer'] = row[2].value
    #         d['symbol'] = row[3].value
    #         d['contract_type'] = row[4].value
    #         d['supply'] = float(row[5].value)
    #         d['base_price'] = row[6].value
    #         d['suppliers_offer'] = row[7].value
    #         d['final_demand'] = row[8].value
    #         d['payan_sabz'] = row[9].value
    #         d['highest_demand_price'] = row[10].value
    #         d['traded_amount'] = row[11].value
    #         d['least_traded_price_rials'] = row[12].value
    #         d['Average_traded_price_rials'] = row[13].value
    #         d['Average_traded_price_origcurrency'] = row[14].value
    #         d['highest_traded_price_rials'] = row[15].value
    #         d['value_KRials'] = row[16].value
    #         d['delivery_date'] = row[17].value
    #         d['subgroup'] = row[18].value
    #         d['group'] = row[19].value
    #         d['main_group'] = row[20].value
    #         d['details'] = row[21].value
    #         d['supplier'] = row[22].value
    #         d['method_of_supply'] = row[23].value
    #         d['supply_code'] = row[24].value
    #         d['supply_type'] = row[25].value
    #
    #         self.datas.append(d)
    #
    # def get(self, request, *args, **kwargs):
    #     per = request.GET['select_and_time[date]']
    #     time_slot = 12
    #     code = {u"۰":"0",u"۱":"1",u"۲":"2",u"۳":"3",u"۴":"4",u"۵":"5",u"۶":"6",u"۷":"7",u"۸":"8",u"۹":"9", u"/":"/"}
    #     new_per = ''.join(code.get(ch, ch) for ch in per)
    #     dates = new_per.split('/')
    #
    #     if (dates == ['']):
    #         persian_date = jdate.date(1394, 8, 1)
    #     else:
    #         persian_date = jdate.date(int(dates[0]), int(dates[1]), int(dates[2]))
    #     result = self.get_product_producer('copper',persian_date, time_slot )
    #     datas = result
    #     output = {'datas': datas}
    #     return HttpResponse(json.dumps(output))


    def initialize(self):
        wb = load_workbook(filename='imev0/test2.xlsx', read_only=True)
        ws = wb.active  # ws is now an IterableWorksheet
        for row in ws:
            d = {}

            if row[0].value == 'تاریخ':
                continue
            dates = row[0].value.split('/')
            persian_date = jdate.date(int(dates[0]), int(dates[1]), int(dates[2]))

            d['date'] = persian_date
            d['name'] = row[1].value
            d['producer'] = row[2].value
            d['symbol'] = row[3].value
            d['contract_type'] = row[4].value
            d['supply'] = float(row[5].value)
            d['base_price'] = row[6].value
            d['suppliers_offer'] = row[7].value
            d['final_demand'] = row[8].value
            d['payan_sabz'] = row[9].value
            d['highest_demand_price'] = row[10].value
            d['traded_amount'] = row[11].value
            d['least_traded_price_rials'] = row[12].value
            d['Average_traded_price_rials'] = row[13].value
            d['Average_traded_price_origcurrency'] = row[14].value
            d['highest_traded_price_rials'] = row[15].value
            d['value_KRials'] = row[16].value
            d['delivery_date'] = row[17].value
            d['subgroup'] = row[18].value
            d['group'] = row[19].value
            d['main_group'] = row[20].value
            d['details'] = row[21].value
            d['supplier'] = row[22].value
            d['method_of_supply'] = row[23].value
            d['supply_code'] = row[24].value
            d['supply_type'] = row[25].value

            self.datas.append(d)

    def set_to_wednesday(self, date):

        while date.weekday() != 4:
            date += jdate.timedelta(days = 1)
        return  date

    # This method takes a symbol which represents a certain product produced by a certain company
    # an end date and a time slot with together represent a time duration. For example if time_slot
    # equals 1, it represents a one week duration ending with end_date (time_slot is given in weeks).

    def string_to_jdate(self, mstring):
        dates = mstring.split("/")
        persian_date = jdate.date(int(dates[0]), int(dates[1]), int(dates[2]))
        return persian_date

    def get(self, request, *args, **kwargs):
        #print("here ", request.GET)
        print('here')
        main_group = request.GET['main_group']
        #print(main_group)
        group = request.GET['group']
        #print(group)
        product = request.GET['product']
        #print(product)
        producer = request.GET['producer']
        #print(producer)
        chart_name = request.GET['type']
        #print(chart_name)
        end_date = request.GET['end_date']
        #print(end_date)
        end_date = self.string_to_jdate(end_date)
        #print(end_date)
        time_slot = (int)(request.GET['time_slot'])
        #print(time_slot)
        sub_group = request.GET['sub_group']
        #print(sub_group)
        data = self.extract_chart(main_group, sub_group, group, product, producer, chart_name, end_date, time_slot)
        output = {'data': data}
        print(output)
        return HttpResponse(json.dumps(output, ensure_ascii=False))


    def extract_chart(self, main_group, sub_group, group, product, producer, chart_name, end_date, time_slot):

        def filter_products():
            trans = Transaction.objects.all()
            if producer != 'all':
                trans = trans.filter(producer__name = producer)
            if main_group != 'all':
                trans = trans.filter(product__group__subGroup__mainGroup__name = main_group)
            if sub_group != 'all':
                trans = trans.filter(product__group__subGroup__name = sub_group)

            if group != 'all':
                trans = trans.filter(product__group__name = group)

            if product != 'all':
                trans = trans.filter(product__name = product)

            return trans




        # the list representing the labels of a line chart
        x = []
        # the list representing the data of a line chart
        y = []
        # the return value (a tuple)
        d = None

        # Now we have the list of all transactions regarding that product
        all_transactions = filter_products()

        if time_slot == ONE_WEEK or time_slot == THREE_WEEK:
        # Considering a week has 7 days, we extract those rows from our database
        # that match the given symbol and time duration given as inputs to the method
            start_date = end_date - timedelta(days = time_slot*7) + timedelta(days = 1)
            date = start_date
            print(end_date)
            while end_date.__ge__(date):
                print(date)
                #TODO what if the result was null? what would it return?
                myDate = datetime.date(date.year, date.month, date.day)
                sum_value = all_transactions.filter(date = myDate).aggregate(Sum(chart_name))[chart_name+'__sum']


                if sum_value is not None and sum_value > 0: #we have data for this particular day!
                    y.append(sum_value)
                    x.append(str(date))
                date += timedelta(days = 1)
            d = (x, y)


        elif time_slot == THREE_MONTHS:
            print("aqa jan three monthse dige, chera shak dari")
            end_date = self.set_to_wednesday(end_date)
            print(end_date)
            start_date = end_date - timedelta(days=12*7)
            print(start_date)
            date = start_date
            while date <= end_date:
                print(date.day , "   " , date.month)
                myDate = datetime.date(date.year, date.month, date.day)
                transactions = all_transactions.filter(date__gt = myDate)

                sum_value = transactions.filter(date__lte =  (myDate + timedelta(days=7))).aggregate(Sum(chart_name))[chart_name+'__sum']
                # debug = transactions.filter(date__lte =  (myDate + timedelta(days=7))).order_by('date', 'product__name')
                # for t in debug:
                #     print(t)
                # print(sum_value)
                if sum_value is not None and sum_value > 0:
                    x.append(str(date + timedelta(days = 7)))
                    y.append(sum_value)
                date += timedelta(days = 7)
            d = (x, y)

        elif time_slot == ONE_YEAR:
            start_date = end_date - timedelta(days = 365)
            # print(str(start_date)+ 'تاریخ شروع ')
            # print(str(end_date) + 'تاریخ پایان ')
            date = start_date
            myDate = datetime.date(date.year, date.month, date.day)

            while date <= end_date:
                # print(str(date) + 'curr date')
                sum_value = all_transactions.filter(date__year = myDate.year).filter(date__month = myDate.month).aggregate(Sum(chart_name))[chart_name+'__sum']

                if sum_value is not None and sum_value >  0:
                    x.append(date.j_months_fa[date.month - 1])
                    y.append(sum_value)
                date += timedelta(days = 30)
            d = (x, y)
        return d


class MainGroupSums(View):
    def get(self, request, *args, **kwargs):
        output = {}

        chart_name = request.GET['type']
        end_date = request.GET['end_date']
        end_date = Datas.string_to_jdate(self,end_date)

        time_slot = (int)(request.GET['time_slot'])
        main_groups = MainGroup.objects.all()
        for item in main_groups:
            trans = Transaction.objects.filter(product__group__subGroup__mainGroup__name = item)
            sum = self.extract_datas(time_slot,end_date,trans,chart_name)
            output[item] = sum
            print(sum)
        print('again in here')
        return HttpResponse(json.dumps(output, ensure_ascii=False))



    def extract_datas(self,time_slot,end_date,all_transactions,chart_name):
        x =[]
        y =[]
        sum = 0
        if time_slot == ONE_WEEK or time_slot == THREE_WEEK:
        # Considering a week has 7 days, we extract those rows from our database
        # that match the given symbol and time duration given as inputs to the method
            start_date = end_date - timedelta(days = time_slot*7) + timedelta(days = 1)
            date = start_date
            print(end_date)
            while end_date.__ge__(date):
                print(date)
                #TODO what if the result was null? what would it return?
                myDate = datetime.date(date.year, date.month, date.day)
                sum_value = all_transactions.filter(date = myDate).aggregate(Sum(chart_name))[chart_name+'__sum']


                if sum_value is not None and sum_value > 0: #we have data for this particular day!
                    y.append(sum_value)
                    x.append(str(date))
                date += timedelta(days = 1)
            d = (x, y)


        elif time_slot == THREE_MONTHS:
            print("aqa jan three monthse dige, chera shak dari")
            end_date = Datas.set_to_wednesday(self,end_date)
            print(end_date)
            start_date = end_date - timedelta(days=12*7)
            print(start_date)
            date = start_date
            while date <= end_date:
                print(date.day , "   " , date.month)
                myDate = datetime.date(date.year, date.month, date.day)
                transactions = all_transactions.filter(date__gt = myDate)

                sum_value = transactions.filter(date__lte =  (myDate + timedelta(days=7))).aggregate(Sum(chart_name))[chart_name+'__sum']
                # debug = transactions.filter(date__lte =  (myDate + timedelta(days=7))).order_by('date', 'product__name')
                # for t in debug:
                #     print(t)
                # print(sum_value)
                if sum_value is not None and sum_value > 0:
                    x.append(str(date + timedelta(days = 7)))
                    y.append(sum_value)
                date += timedelta(days = 7)
            d = (x, y)

        elif time_slot == ONE_YEAR:
            start_date = end_date - timedelta(days = 365)
            # print(str(start_date)+ 'تاریخ شروع ')
            # print(str(end_date) + 'تاریخ پایان ')
            date = start_date
            myDate = datetime.date(date.year, date.month, date.day)

            while date <= end_date:
                # print(str(date) + 'curr date')
                sum_value = all_transactions.filter(date__year = myDate.year).filter(date__month = myDate.month).aggregate(Sum(chart_name))[chart_name+'__sum']

                if sum_value is not None and sum_value >  0:
                    x.append(date.j_months_fa[date.month - 1])
                    y.append(sum_value)
                date += timedelta(days = 30)
            d = (x, y)


        print('pfff')
        for item in y:
            sum = sum + item
        return sum
